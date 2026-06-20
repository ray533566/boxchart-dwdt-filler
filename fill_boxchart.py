"""
fill_boxchart.py
----------------
Fills Format_Mode_hopping_BoxChart620.xlsx with aggregated dW/dT data
from RawData_for_BoxPlot.xlsx.

Mapping:
  RawData CHNumber          -> Sheet                        -> CH
  1_Operational             -> Normal_Operational Current   -> CH1
  2_Operational             -> Normal_Operational Current   -> CH2
  3_Operational             -> Normal_Operational Current   -> CH3
  4_Operational             -> Normal_Operational Current   -> CH4
  1_Maximum                 -> Bias400_Maximum Current      -> CH1
  2_Maximum                 -> Bias400_Maximum Current      -> CH2
  3_Maximum                 -> Bias400_Maximum Current      -> CH3
  4_Maximum                 -> Bias400_Maximum Current      -> CH4

Column layout per sheet:
  A        : TESTSN
  B-M      : avg / max / min of dW/dT  x  CH1, CH2, CH3, CH4
  N, O     : empty (preserved as-is)
  P-AA     : max / avg / min of dW/dT  x  CH1, CH2, CH3, CH4

Box plot charts are preserved by patching the sheet XML directly
inside the ZIP archive (openpyxl would strip chartEx objects).

Usage:
    python fill_boxchart.py
    python fill_boxchart.py --raw RawData.xlsx --fmt Format.xlsx --out Output.xlsx
"""

import argparse
import re
import shutil
import zipfile
from pathlib import Path

import pandas as pd


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

DEFAULT_RAW  = "RawData_for_BoxPlot.xlsx"
DEFAULT_FMT  = "Format_Mode_hopping_BoxChart620.xlsx"
DEFAULT_OUT  = "Format_Mode_hopping_BoxChart620_updated.xlsx"

OP_CHANNELS = ["1_Operational", "2_Operational", "3_Operational", "4_Operational"]
MX_CHANNELS = ["1_Maximum",     "2_Maximum",     "3_Maximum",     "4_Maximum"]

# sheet name -> (internal sheet file, channel list)
SHEET_MAP = {
    "Normal_Operational Current": ("xl/worksheets/sheet2.xml", OP_CHANNELS),
    "Bias400_Maximum Current":    ("xl/worksheets/sheet3.xml", MX_CHANNELS),
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def col_letter(n: int) -> str:
    """Convert 1-based column index to Excel letter(s). e.g. 1->A, 27->AA"""
    if n <= 26:
        return chr(64 + n)
    return chr(64 + (n - 1) // 26) + chr(64 + (n - 1) % 26 + 1)


def agg_dwdt(df: pd.DataFrame, sn: str, channels: list) -> list:
    """
    For each channel, compute avg / max / min of dW/dT for the given SN.
    Returns a flat list: [avg1, max1, min1, avg2, max2, min2, ...]
    """
    result = []
    for ch in channels:
        sub = df.loc[(df["TESTSN"] == sn) & (df["CHNumber"] == ch), "dW/dT"]
        if sub.empty:
            result.extend([None, None, None])
        else:
            result.extend([
                float(round(sub.mean(), 4)),
                float(round(sub.max(),  4)),
                float(round(sub.min(),  4)),
            ])
    return result


def parse_shared_strings(xml_bytes: bytes) -> list:
    """Return list of string values from xl/sharedStrings.xml."""
    ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    from lxml import etree
    tree = etree.fromstring(xml_bytes)
    strings = []
    for si in tree.findall(f"{{{ns}}}si"):
        t = si.find(f".//{{{ns}}}t")
        strings.append(t.text if t is not None else "")
    return strings


def build_shared_strings_xml(strings: list) -> bytes:
    """Rebuild xl/sharedStrings.xml from a list of string values."""
    ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    count = len(strings)
    # Escape XML special characters
    def esc(s):
        return (s.replace("&", "&amp;")
                  .replace("<", "&lt;")
                  .replace(">", "&gt;")
                  .replace('"', "&quot;"))
    items = "".join(f"<si><t>{esc(s)}</t></si>" for s in strings)
    xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<sst xmlns="{ns}" count="{count}" uniqueCount="{count}">{items}</sst>'
    )
    return xml.encode("utf-8")


def build_data_rows_xml(
    data_rows: list,
    shared_strings: list,
    start_row: int = 2,
) -> tuple:
    """
    Build a list of <row> XML strings for the data section.

    Column layout written:
      A       (col 1)  : SN  (shared string)
      B-M     (col 2-13): avg, max, min per CH1-4
      N, O    (col 14-15): skipped
      P-AA    (col 16-27): max, avg, min per CH1-4

    Returns (list_of_row_xml_strings, updated_shared_strings).
    """
    ss_map = {v: i for i, v in enumerate(shared_strings)}
    new_shared = list(shared_strings)

    rows_xml = []
    for i, row_data in enumerate(data_rows):
        excel_row = start_row + i
        sn = row_data[0]

        # Ensure SN is in shared strings
        if sn not in ss_map:
            ss_map[sn] = len(new_shared)
            new_shared.append(sn)

        cells = []

        # --- Column A: SN (string) ---
        cells.append(f'<c r="A{excel_row}" t="s"><v>{ss_map[sn]}</v></c>')

        # --- Columns B-M: avg, max, min x CH1-4 ---
        # row_data[1..12] = [avg1, max1, min1, avg2, max2, min2, avg3, max3, min3, avg4, max4, min4]
        for offset in range(12):           # 0..11
            col = offset + 2               # 2=B .. 13=M
            val = row_data[1 + offset]
            if val is not None:
                cells.append(f'<c r="{col_letter(col)}{excel_row}"><v>{val}</v></c>')

        # --- Columns P-AA: max, avg, min x CH1-4 ---
        # Reorder from B-M layout (avg/max/min) to P-AA layout (max/avg/min)
        # CH1: B=avg1(idx1), C=max1(idx2), D=min1(idx3)  -> P=max, Q=avg, R=min
        # CH2: E=avg2(idx4), F=max2(idx5), G=min2(idx6)  -> S=max, T=avg, U=min
        # CH3: H=avg3(idx7), I=max3(idx8), J=min3(idx9)  -> V=max, W=avg, X=min
        # CH4: K=avg4(idx10),L=max4(idx11),M=min4(idx12) -> Y=max, Z=avg, AA=min
        p_aa_indices = [2, 1, 3,  5, 4, 6,  8, 7, 9,  11, 10, 12]
        for offset, val_idx in enumerate(p_aa_indices):
            col = offset + 16              # 16=P .. 27=AA
            val = row_data[val_idx]
            if val is not None:
                cells.append(f'<c r="{col_letter(col)}{excel_row}"><v>{val}</v></c>')

        rows_xml.append(f'<row r="{excel_row}" spans="1:27">{"".join(cells)}</row>')

    return rows_xml, new_shared


def replace_sheet_data(sheet_xml: str, new_rows_xml: list) -> str:
    """
    Replace all data rows (row >= 2) inside <sheetData> while keeping the
    header row (row 1) and all other sheet elements untouched.
    """
    sd_start = sheet_xml.index("<sheetData>")
    sd_end   = sheet_xml.index("</sheetData>") + len("</sheetData>")
    sheet_data_content = sheet_xml[sd_start:sd_end]

    row1_match = re.search(r'<row r="1"[^>]*>.*?</row>', sheet_data_content, re.DOTALL)
    row1_xml   = row1_match.group(0) if row1_match else ""

    new_sheet_data = f"<sheetData>{row1_xml}{''.join(new_rows_xml)}</sheetData>"
    return sheet_xml[:sd_start] + new_sheet_data + sheet_xml[sd_end:]


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main(raw_path: str, fmt_path: str, out_path: str) -> None:
    print(f"Reading raw data : {raw_path}")
    df = pd.read_excel(raw_path)
    required_cols = {"TESTSN", "CHNumber", "dW/dT"}
    if not required_cols.issubset(df.columns):
        raise ValueError(f"RawData missing columns: {required_cols - set(df.columns)}")

    raw_sns = sorted(df["TESTSN"].unique().tolist())
    print(f"  Unique TESTSN  : {len(raw_sns)}")

    # -----------------------------------------------------------------------
    # Build aggregated tables
    # -----------------------------------------------------------------------
    print("Aggregating dW/dT (avg / max / min) per SN and channel …")

    op_rows = [[sn] + agg_dwdt(df, sn, OP_CHANNELS) for sn in raw_sns]
    mx_rows = [[sn] + agg_dwdt(df, sn, MX_CHANNELS) for sn in raw_sns]

    # -----------------------------------------------------------------------
    # Read original ZIP contents
    # -----------------------------------------------------------------------
    print(f"Reading format file: {fmt_path}")
    with zipfile.ZipFile(fmt_path, "r") as zin:
        orig_contents = {name: zin.read(name) for name in zin.namelist()}

    shared_strings = parse_shared_strings(orig_contents["xl/sharedStrings.xml"])

    # -----------------------------------------------------------------------
    # Patch sheet XMLs
    # -----------------------------------------------------------------------
    sheet_file_op = "xl/worksheets/sheet2.xml"  # Normal_Operational Current
    sheet_file_mx = "xl/worksheets/sheet3.xml"  # Bias400_Maximum Current

    op_row_xmls, shared_after_op = build_data_rows_xml(op_rows, shared_strings, start_row=2)
    mx_row_xmls, shared_final    = build_data_rows_xml(mx_rows, shared_after_op, start_row=2)

    new_sheet2 = replace_sheet_data(orig_contents[sheet_file_op].decode("utf-8"), op_row_xmls)
    new_sheet3 = replace_sheet_data(orig_contents[sheet_file_mx].decode("utf-8"), mx_row_xmls)

    new_ss_xml = build_shared_strings_xml(shared_final)

    # -----------------------------------------------------------------------
    # Write output ZIP (preserves all chart files)
    # -----------------------------------------------------------------------
    print(f"Writing output   : {out_path}")
    replacements = {
        sheet_file_op:            new_sheet2.encode("utf-8"),
        sheet_file_mx:            new_sheet3.encode("utf-8"),
        "xl/sharedStrings.xml":   new_ss_xml,
    }

    with zipfile.ZipFile(out_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for name, content in orig_contents.items():
            zout.writestr(name, replacements.get(name, content))

    # -----------------------------------------------------------------------
    # Verification
    # -----------------------------------------------------------------------
    print("\nVerification (first 3 SNs, CH1, Normal_Operational Current):")
    from openpyxl import load_workbook
    wb  = load_workbook(out_path, data_only=True)
    ws2 = wb["Normal_Operational Current"]
    ws3 = wb["Bias400_Maximum Current"]

    all_ok = True
    for i, sn in enumerate(raw_sns[:3]):
        row = i + 2
        sub = df.loc[(df["TESTSN"] == sn) & (df["CHNumber"] == "1_Operational"), "dW/dT"]
        exp_avg = round(sub.mean(), 4)
        exp_max = round(sub.max(),  4)
        exp_min = round(sub.min(),  4)

        got_avg = ws2.cell(row, 2).value   # B = avg
        got_max = ws2.cell(row, 3).value   # C = max
        got_min = ws2.cell(row, 4).value   # D = min
        got_p   = ws2.cell(row, 16).value  # P = max (reordered)
        got_q   = ws2.cell(row, 17).value  # Q = avg
        got_r   = ws2.cell(row, 18).value  # R = min

        bm_ok  = abs(got_avg - exp_avg) < 1e-6 and abs(got_max - exp_max) < 1e-6 and abs(got_min - exp_min) < 1e-6
        paa_ok = abs(got_p - exp_max) < 1e-6   and abs(got_q - exp_avg) < 1e-6   and abs(got_r - exp_min) < 1e-6
        status = "OK" if (bm_ok and paa_ok) else "FAIL"
        if not (bm_ok and paa_ok):
            all_ok = False
        print(f"  {sn}  B-M:{status}  P-AA:{status}  avg={got_avg}  max={got_max}  min={got_min}")

    with zipfile.ZipFile(out_path) as zcheck:
        chart_count = sum(1 for f in zcheck.namelist() if "chart" in f)
    print(f"\nChart files preserved : {chart_count}")
    print(f"Op  sheet rows        : {ws2.max_row - 1}")
    print(f"Max sheet rows        : {ws3.max_row - 1}")
    print(f"\n{'All checks passed.' if all_ok else 'WARNING: some checks failed!'}")
    print(f"Output saved to: {out_path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Fill BoxChart format file with dW/dT aggregations.")
    parser.add_argument("--raw", default=DEFAULT_RAW, help="Path to RawData_for_BoxPlot.xlsx")
    parser.add_argument("--fmt", default=DEFAULT_FMT, help="Path to Format_Mode_hopping_BoxChart620.xlsx")
    parser.add_argument("--out", default=DEFAULT_OUT, help="Output file path")
    args = parser.parse_args()

    main(args.raw, args.fmt, args.out)
